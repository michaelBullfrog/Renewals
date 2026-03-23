#!/usr/bin/env python3
"""Render-ready renewal notifier for Webex.

Reads an Excel workbook, keeps ACTIVE subscriptions only, finds subscriptions
whose renewal date is exactly N days from today, sorts soonest-first, and posts
those rows to a Webex space.

Environment variables:
  WEBEX_ACCESS_TOKEN   required
  WEBEX_ROOM_ID        required
  EXCEL_FILE           default: subscriptions.xlsx
  SHEET_NAME           default: Subscription Details-Line Level
  STATUS_FILTER        default: ACTIVE
  NOTIFY_DAYS          default: 60
  PER_MESSAGE          default: 15
  INCLUDE_HEADER       default: true
  ALLOW_DUPLICATES     default: false
  STATE_FILE           optional path to JSON state file used to avoid duplicate sends
  RUN_DATE             optional YYYY-MM-DD override for testing

Example local run:
  WEBEX_ACCESS_TOKEN=... WEBEX_ROOM_ID=... python render_renewal_notifier.py
"""
from __future__ import annotations

import json
import os
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Iterable

import openpyxl
import requests

WEBEX_URL = "https://webexapis.com/v1/messages"
REQUIRED_COLUMNS = ["End Customer", "Subscription ID", "Renewal Date", "Status"]
DATE_FORMATS = [
    "%d %b %Y",
    "%m/%d/%Y",
    "%Y-%m-%d",
    "%m-%d-%Y",
    "%d-%b-%Y",
    "%B %d, %Y",
    "%b %d, %Y",
]


@dataclass
class Config:
    token: str
    room_id: str
    excel_file: str = "subscriptions.xlsx"
    sheet_name: str = "Subscription Details-Line Level"
    status_filter: str = "ACTIVE"
    notify_days: int = 60
    per_message: int = 15
    include_header: bool = True
    allow_duplicates: bool = False
    state_file: str | None = None
    run_date: date | None = None


@dataclass
class RenewalRow:
    end_customer: str
    subscription_id: str
    renewal_date_text: str
    renewal_date: date
    status: str

    @property
    def dedupe_key(self) -> str:
        return f"{self.end_customer}|{self.subscription_id}|{self.renewal_date.isoformat()}|{self.status}"


def env_bool(name: str, default: bool) -> bool:
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "y", "on"}


def parse_date(value) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def format_date(value) -> str:
    parsed = parse_date(value)
    if parsed is None:
        return ""
    return parsed.strftime("%d %b %Y")


def chunked(items: list[RenewalRow], size: int) -> Iterable[list[RenewalRow]]:
    for i in range(0, len(items), size):
        yield items[i : i + size]


def get_config() -> Config:
    token = os.getenv("WEBEX_ACCESS_TOKEN", "").strip()
    room_id = os.getenv("WEBEX_ROOM_ID", "").strip()
    if not token:
        raise SystemExit("Missing WEBEX_ACCESS_TOKEN environment variable.")
    if not room_id:
        raise SystemExit("Missing WEBEX_ROOM_ID environment variable.")

    run_date_text = os.getenv("RUN_DATE", "").strip()
    run_date = None
    if run_date_text:
        run_date = datetime.strptime(run_date_text, "%Y-%m-%d").date()

    return Config(
        token=token,
        room_id=room_id,
        excel_file=os.getenv("EXCEL_FILE", "subscriptions.xlsx").strip(),
        sheet_name=os.getenv("SHEET_NAME", "Subscription Details-Line Level").strip(),
        status_filter=os.getenv("STATUS_FILTER", "ACTIVE").strip().upper(),
        notify_days=int(os.getenv("NOTIFY_DAYS", "60")),
        per_message=max(1, int(os.getenv("PER_MESSAGE", "15"))),
        include_header=env_bool("INCLUDE_HEADER", True),
        allow_duplicates=env_bool("ALLOW_DUPLICATES", False),
        state_file=os.getenv("STATE_FILE", "").strip() or None,
        run_date=run_date,
    )


def find_columns(ws) -> dict[str, int]:
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    lookup: dict[str, int] = {}
    for idx, header in enumerate(headers, start=1):
        if header is not None:
            lookup[str(header).strip()] = idx

    missing = [name for name in REQUIRED_COLUMNS if name not in lookup]
    if missing:
        raise ValueError(
            f"Missing required column(s): {', '.join(missing)}. Found headers include: {headers[:25]}"
        )
    return {name: lookup[name] for name in REQUIRED_COLUMNS}


def load_candidate_rows(cfg: Config) -> list[RenewalRow]:
    workbook = openpyxl.load_workbook(cfg.excel_file, data_only=True)
    if cfg.sheet_name not in workbook.sheetnames:
        raise ValueError(
            f"Sheet '{cfg.sheet_name}' not found. Available sheets: {', '.join(workbook.sheetnames)}"
        )

    ws = workbook[cfg.sheet_name]
    col = find_columns(ws)
    target_date = (cfg.run_date or date.today()) + timedelta(days=cfg.notify_days)
    rows: list[RenewalRow] = []
    seen: set[str] = set()

    for r in range(2, ws.max_row + 1):
        status_raw = ws.cell(r, col["Status"]).value
        status = "" if status_raw is None else str(status_raw).strip().upper()
        if status != cfg.status_filter:
            continue

        renewal_raw = ws.cell(r, col["Renewal Date"]).value
        renewal_date = parse_date(renewal_raw)
        if renewal_date is None or renewal_date != target_date:
            continue

        end_customer_raw = ws.cell(r, col["End Customer"]).value
        subscription_id_raw = ws.cell(r, col["Subscription ID"]).value

        row = RenewalRow(
            end_customer="" if end_customer_raw is None else str(end_customer_raw).strip(),
            subscription_id="" if subscription_id_raw is None else str(subscription_id_raw).strip(),
            renewal_date_text=format_date(renewal_raw),
            renewal_date=renewal_date,
            status=status,
        )

        if not (row.end_customer or row.subscription_id or row.renewal_date_text):
            continue

        if not cfg.allow_duplicates and row.dedupe_key in seen:
            continue
        seen.add(row.dedupe_key)
        rows.append(row)

    rows.sort(key=lambda row: (row.renewal_date, row.end_customer, row.subscription_id))
    return rows


def load_state(path: str | None) -> dict:
    if not path:
        return {"sent_keys": []}
    state_path = Path(path)
    if not state_path.exists():
        return {"sent_keys": []}
    try:
        return json.loads(state_path.read_text(encoding="utf-8"))
    except Exception:
        return {"sent_keys": []}


def save_state(path: str | None, state: dict) -> None:
    if not path:
        return
    state_path = Path(path)
    state_path.parent.mkdir(parents=True, exist_ok=True)
    state_path.write_text(json.dumps(state, indent=2, sort_keys=True), encoding="utf-8")


def filter_unsent(rows: list[RenewalRow], state: dict) -> list[RenewalRow]:
    sent = set(state.get("sent_keys", []))
    return [row for row in rows if row.dedupe_key not in sent]


def mark_sent(rows: list[RenewalRow], state: dict) -> dict:
    sent = set(state.get("sent_keys", []))
    for row in rows:
        sent.add(row.dedupe_key)
    state["sent_keys"] = sorted(sent)
    state["last_updated_utc"] = datetime.utcnow().isoformat() + "Z"
    return state


def build_markdown(rows: list[RenewalRow], cfg: Config, part: int, total_parts: int) -> str:
    lines: list[str] = []
    target_date = (cfg.run_date or date.today()) + timedelta(days=cfg.notify_days)
    if cfg.include_header:
        header = (
            f"**Subscription renewals due in {cfg.notify_days} days**"
            f" — target date: {target_date.strftime('%d %b %Y')}"
        )
        if total_parts > 1:
            header += f" ({part}/{total_parts})"
        lines.extend([header, ""])

    for row in rows:
        lines.append(
            f"- **End Customer:** {row.end_customer}  \n"
            f"  **Subscription ID:** {row.subscription_id}  \n"
            f"  **Renewal Date:** {row.renewal_date_text}"
        )
    return "\n".join(lines)


def post_message(cfg: Config, markdown: str) -> None:
    response = requests.post(
        WEBEX_URL,
        headers={
            "Authorization": f"Bearer {cfg.token}",
            "Content-Type": "application/json",
        },
        json={"roomId": cfg.room_id, "markdown": markdown},
        timeout=30,
    )
    response.raise_for_status()


def main() -> None:
    cfg = get_config()
    state = load_state(cfg.state_file)
    all_rows = load_candidate_rows(cfg)
    rows_to_send = filter_unsent(all_rows, state)

    if not rows_to_send:
        target_date = (cfg.run_date or date.today()) + timedelta(days=cfg.notify_days)
        print(
            f"No unsent {cfg.status_filter} renewals found for {target_date.isoformat()} "
            f"({cfg.notify_days} days out)."
        )
        return

    chunks = list(chunked(rows_to_send, cfg.per_message))
    for index, chunk in enumerate(chunks, start=1):
        markdown = build_markdown(chunk, cfg, index, len(chunks))
        post_message(cfg, markdown)

    state = mark_sent(rows_to_send, state)
    save_state(cfg.state_file, state)
    print(
        f"Sent {len(rows_to_send)} renewal notification row(s) in {len(chunks)} Webex message(s)."
    )


if __name__ == "__main__":
    main()
